"""
Vicio Promo Tracker — Weekly Excel Updater
==========================================
Cada lunes:
  1. Descarga el Excel de Google Drive (o usa el local)
  2. Lee las filas de la semana anterior
  3. Para cada partner/plataforma, busca con Claude (web search):
       - DF, SF, MBS actuales
       - Promos activas (Comments)
  4. Escribe las nuevas filas con el número de semana correcto
  5. Envía el Excel actualizado por Gmail
"""

import os, json, re, smtplib, shutil, datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import openpyxl
import anthropic

# ─── CONFIG (se leen desde variables de entorno de GitHub Actions) ─────────────
ANTHROPIC_KEY   = os.environ["ANTHROPIC_API_KEY"]
EMAIL_REMITENTE = os.environ["EMAIL_REMITENTE"]
EMAIL_DESTINO   = os.environ["EMAIL_DESTINO"]
GMAIL_APP_PASS  = os.environ["GMAIL_APP_PASS"]
EXCEL_PATH      = os.environ.get("EXCEL_PATH", ""XL ES Glossary - group & brand .xlsx"")
# ──────────────────────────────────────────────────────────────────────────────

SHEET_NAME = "Competitors promos & pricing"

PARTNERS = [
    "McD", "Telepizza", "KFC", "Papa John's", "BK", "Popeyes",
    "Domino's Pizza", "Goiko", "VIPS", "Taco Bell",
    "Foster's Hollywood", "Pizzeria Carlos", "Five Guys"
]

PLATFORMS = ["Glovo", "UberEats", "JustEat"]

# Nombre completo para búsquedas web
PARTNER_FULL_NAME = {
    "McD":               "McDonald's España",
    "Telepizza":         "Telepizza España",
    "KFC":               "KFC España",
    "Papa John's":       "Papa John's España",
    "BK":                "Burger King España",
    "Popeyes":           "Popeyes España",
    "Domino's Pizza":    "Domino's Pizza España",
    "Goiko":             "Goiko España",
    "VIPS":              "VIPS España",
    "Taco Bell":         "Taco Bell España",
    "Foster's Hollywood":"Foster's Hollywood España",
    "Pizzeria Carlos":   "Pizzeria Carlos España",
    "Five Guys":         "Five Guys España",
}

PLATFORM_SEARCH_NAME = {
    "Glovo":    "Glovo",
    "UberEats": "Uber Eats",
    "JustEat":  "Just Eat",
}


def get_current_week() -> tuple[int, str]:
    """Devuelve (año, 'wNN') de la semana actual."""
    today = datetime.date.today()
    week  = today.isocalendar()[1]
    return today.year, f"w{week}"


def get_last_week_data(ws) -> dict:
    """
    Lee las filas más recientes del Excel y devuelve un dict:
    {partner: {platform: {owner, am, df, sf, mbs, df_promo, promo_in_menu,
                           promocode, proper_delivery, comments}}}
    """
    # Encontrar la semana más reciente
    latest_year, latest_week = 0, ""
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] and row[1]:
            yr = int(row[0])
            wk = str(row[1]).lower().strip()
            wk_num = int(re.sub(r'\D', '', wk) or 0)
            if yr > latest_year or (yr == latest_year and wk_num > int(re.sub(r'\D', '', latest_week) or 0)):
                latest_year  = yr
                latest_week  = wk

    print(f"📅 Semana más reciente en Excel: {latest_year} {latest_week}")

    data = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0] or not row[1]:
            continue
        if int(row[0]) == latest_year and str(row[1]).lower().strip() == latest_week:
            partner  = str(row[4]).strip() if row[4] else ""
            platform = str(row[5]).strip() if row[5] else ""
            if partner not in PARTNERS or platform not in PLATFORMS:
                continue
            if partner not in data:
                data[partner] = {}
            data[partner][platform] = {
                "owner":            row[2],
                "am":               row[3],
                "df":               row[6],
                "sf":               row[7],
                "mbs":              row[8],
                "df_promo":         row[9],
                "promo_in_menu":    row[10],
                "promocode":        row[11],
                "proper_delivery":  row[12],
                "comments":         row[13],
            }
    return data


def search_partner_data(partner: str, platform: str) -> dict:
    """
    Usa Claude con web_search para buscar DF, SF, MBS y promos actuales
    de un partner en una plataforma. Devuelve un dict con los campos.
    """
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    full_name     = PARTNER_FULL_NAME.get(partner, partner)
    platform_name = PLATFORM_SEARCH_NAME.get(platform, platform)

    prompt = f"""Busca en {platform_name} España los datos actuales de {full_name} hoy mismo.
Necesito exactamente estos datos en formato JSON (sin texto extra, sin backticks):
{{
  "df": "coste de entrega (Delivery Fee) en euros, ej: 1,99€ o 0€",
  "sf": "service fee o comisión de servicio, ej: 0,75€ o 8% del pedido",
  "mbs": "pedido mínimo (Minimum Basket Size) si existe, ej: If < 15€ surcharge 2€ o No MBS",
  "comments": "lista de promociones activas ahora mismo, cada una en una línea, con porcentaje y nombre del producto. Si no hay promos escribe: Sin promo"
}}
Sé concreto con los valores reales que aparecen en la app/web. Si no encuentras dato exacto escribe null."""

    try:
        response = client.messages.create(
            model="claude-opus-4-5",
            max_tokens=800,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": prompt}]
        )
        text = "".join(b.text for b in response.content if hasattr(b, "text"))
        text = text.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        m = re.search(r'\{[\s\S]*\}', text)
        if m:
            return json.loads(m.group())
    except Exception as e:
        print(f"  ⚠️  Error buscando {partner}/{platform}: {e}")
    return {"df": None, "sf": None, "mbs": None, "comments": None}


def append_new_week(ws, prev_data: dict, new_year: int, new_week: str):
    """Añade las filas de la nueva semana al final del sheet."""
    next_row = ws.max_row + 1
    # Buscar última fila real con datos
    for r in range(ws.max_row, 3, -1):
        if any(ws.cell(r, c).value for c in range(1, 15)):
            next_row = r + 1
            break

    for partner in PARTNERS:
        platforms_to_write = PLATFORMS
        # Añadir "Own channel" si el partner lo tenía
        partner_prev = prev_data.get(partner, {})

        for platform in platforms_to_write:
            prev = partner_prev.get(platform, {})
            owner = prev.get("owner") or ""
            am    = prev.get("am")    or ""

            print(f"  🔍 Buscando {partner} / {platform}...")
            new_data = search_partner_data(partner, platform)

            df       = new_data.get("df")       or prev.get("df")
            sf       = new_data.get("sf")       or prev.get("sf")
            mbs      = new_data.get("mbs")      or prev.get("mbs")
            comments = new_data.get("comments") or ""

            # Mantener campos que no cambian semanalmente
            df_promo        = prev.get("df_promo")
            promo_in_menu   = prev.get("promo_in_menu")
            promocode       = prev.get("promocode")
            proper_delivery = prev.get("proper_delivery")

            row_data = [
                new_year,       # A: Year
                new_week,       # B: Week
                owner,          # C: Owner
                am,             # D: AM
                partner,        # E: Partner
                platform,       # F: Company
                df,             # G: DF
                sf,             # H: SF
                mbs,            # I: MBS
                df_promo,       # J: DF Promo
                promo_in_menu,  # K: Promo in menu
                promocode,      # L: Promocode (CRM)
                proper_delivery,# M: Proper Delivery
                comments,       # N: Comments
            ]

            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=next_row, column=col_idx, value=value)

            next_row += 1

    print(f"✅ Escritas filas hasta fila {next_row - 1}")


def send_email_with_attachment(filepath: str, new_week: str, new_year: int):
    today    = datetime.date.today().strftime("%d/%m/%Y")
    filename = os.path.basename(filepath)
    subject  = f"[WBR] Excel Competitors Promos actualizado — {new_week} {new_year}"

    body_html = f"""
    <html><body style="font-family:Arial,sans-serif;color:#1a1a1a;max-width:600px;margin:auto">
    <div style="background:#1a1a1a;padding:20px 24px;border-radius:8px 8px 0 0">
      <h1 style="color:#fff;margin:0;font-size:18px">📊 Competitors Promos — Excel actualizado</h1>
      <p style="color:#aaa;margin:4px 0 0;font-size:13px">{new_week.upper()} {new_year} · {today}</p>
    </div>
    <div style="border:1px solid #e5e5e5;border-top:none;padding:24px;border-radius:0 0 8px 8px">
      <p style="font-size:14px">El Excel de seguimiento de competidores ha sido actualizado automáticamente con los datos de la semana actual.</p>
      <ul style="font-size:13px;color:#555;line-height:2">
        <li>✅ DF, SF y MBS actualizados para todos los partners</li>
        <li>✅ Promociones activas (Comments) actualizadas</li>
        <li>📎 Excel adjunto — sube a Google Drive para sustituir el anterior</li>
      </ul>
      <p style="font-size:11px;color:#aaa;margin-top:24px">Generado automáticamente · Vicio Promo Tracker · GitHub Actions</p>
    </div>
    </body></html>
    """

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"]    = EMAIL_REMITENTE
    msg["To"]      = EMAIL_DESTINO
    msg.attach(MIMEText(body_html, "html"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(EMAIL_REMITENTE, GMAIL_APP_PASS)
        server.sendmail(EMAIL_REMITENTE, EMAIL_DESTINO, msg.as_string())

    print(f"✅ Email enviado con adjunto a {EMAIL_DESTINO}")


def main():
    print("=" * 60)
    print("🚀 Vicio Promo Tracker — Actualización semanal")
    print("=" * 60)

    # 1. Cargar Excel
    print(f"\n📂 Cargando Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 2. Leer semana anterior
    print("\n📖 Leyendo datos de la semana anterior...")
    prev_data = get_last_week_data(ws)
    print(f"   Partners encontrados: {list(prev_data.keys())}")

    # 3. Calcular nueva semana
    new_year, new_week = get_current_week()
    print(f"\n📅 Nueva semana: {new_year} {new_week}")

    # 4. Buscar y escribir nuevos datos
    print("\n🔍 Buscando datos actualizados para cada partner/plataforma...")
    append_new_week(ws, prev_data, new_year, new_week)

    # 5. Guardar Excel
    output_path = f"XL_ES_Glossary_{new_week}_{new_year}.xlsx"
    wb.save(output_path)
    print(f"\n💾 Excel guardado: {output_path}")

    # 6. Enviar por email
    print("\n📧 Enviando email con adjunto...")
    send_email_with_attachment(output_path, new_week, new_year)

    print("\n✅ ¡Todo listo! El Excel de esta semana está en tu bandeja de entrada.")


if __name__ == "__main__":
    main()
